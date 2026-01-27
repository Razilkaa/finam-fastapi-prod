"""Excel document generation service."""
from datetime import date
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.date_utils import (
    format_date_ru,
    format_date_en,
    format_sheet_name_ru,
    format_sheet_name_en,
    group_items_by_date,
    choose_reference_monday,
    get_week_dates,
    parse_time_for_sort,
    format_time_display,
)
from app.utils.text_utils import sanitize_text, convert_month_suffix_to_ru
from app.utils.constants import (
    FONT_HEADER,
    FONT_DATE,
    FONT_TIME,
    FONT_TIME_RED,
    FONT_EVENT,
    FONT_EVENT_RED,
    FONT_HOLIDAY,
    GRAY_FILL,
    WHITE_FILL,
    NO_FILL,
    COUNTRY_NAMES_EN,
    COUNTRY_NAMES_RU,
    ALIGN_LEFT,
    ALIGN_CENTER,
    BORDER_THIN,
    BORDER_MEDIUM,
)


class ColumnWidthTracker:
    """Отслеживает максимальную ширину для каждой колонки."""
    def __init__(self):
        self.widths: dict[int, float] = {}
    
    def update(self, col: int, value: str):
        if value:
            text = str(value)
            width = 0.0
            for char in text:
                if ord(char) > 127:
                    width += 1.3
                else:
                    width += 1.0
            self.widths[col] = max(self.widths.get(col, 0), width)
    
    def apply(self, ws: Worksheet, padding: float = 2.0):
        for col, width in self.widths.items():
            col_letter = ws.cell(1, col).column_letter
            ws.column_dimensions[col_letter].width = width + padding


def write_header_row(ws: Worksheet, lang: str, tracker: ColumnWidthTracker):
    """Записывает строку заголовков."""
    if lang == "en":
        headers = ["Date/time", "Country", "News"]
    else:
        headers = ["Дата/Время", "Страна", "Событие"]
    
    for i, header in enumerate(headers):
        col = 3 + i
        cell = ws.cell(2, col)
        cell.value = header
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = Border(
            left=BORDER_MEDIUM if i == 0 else BORDER_THIN,
            right=BORDER_MEDIUM if i == 2 else BORDER_THIN,
            top=BORDER_MEDIUM,
            bottom=BORDER_MEDIUM
        )
        tracker.update(col, header)


def write_date_row(ws: Worksheet, row: int, date_text: str, is_first: bool, tracker: ColumnWidthTracker):
    """Запись строки с датой (объединённые ячейки C:E)."""
    ws.merge_cells(f"C{row}:E{row}")
    cell = ws.cell(row, 3)
    cell.value = date_text
    cell.font = FONT_DATE
    cell.alignment = ALIGN_LEFT
    cell.fill = GRAY_FILL
    top = BORDER_MEDIUM if is_first else BORDER_THIN
    cell.border = Border(left=BORDER_MEDIUM, right=BORDER_MEDIUM, top=top, bottom=BORDER_THIN)

    cell_d = ws.cell(row, 4)
    cell_d.fill = GRAY_FILL
    cell_d.border = Border(top=top, bottom=BORDER_THIN)

    cell_e = ws.cell(row, 5)
    cell_e.fill = GRAY_FILL
    cell_e.border = Border(right=BORDER_MEDIUM, top=top, bottom=BORDER_THIN)

    tracker.update(5, date_text)


def write_event_row(ws: Worksheet, row: int, time_str, country, event,
                    is_important: bool, is_last: bool, tracker: ColumnWidthTracker):
    """Запись строки события."""
    cell_c = ws.cell(row, 3)
    cell_d = ws.cell(row, 4)
    cell_e = ws.cell(row, 5)
    
    time_display = format_time_display(time_str)
    country_safe = sanitize_text(country)
    event_safe = sanitize_text(event)
    
    cell_c.value = time_display
    cell_d.value = country_safe
    cell_e.value = event_safe
    
    tracker.update(3, time_display)
    tracker.update(4, country_safe)
    tracker.update(5, event_safe)
    
    if is_important:
        cell_c.font = FONT_TIME_RED
        cell_d.font = FONT_EVENT_RED
        cell_e.font = FONT_EVENT_RED
    else:
        cell_c.font = FONT_TIME
        cell_d.font = FONT_EVENT
        cell_e.font = FONT_EVENT
    
    cell_c.alignment = ALIGN_LEFT
    cell_d.alignment = ALIGN_LEFT
    cell_e.alignment = ALIGN_LEFT
    cell_c.fill = WHITE_FILL
    cell_d.fill = WHITE_FILL
    cell_e.fill = WHITE_FILL
    
    bottom = BORDER_MEDIUM if is_last else BORDER_THIN
    cell_c.border = Border(left=BORDER_MEDIUM, right=BORDER_THIN, top=BORDER_THIN, bottom=bottom)
    cell_d.border = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=bottom)
    cell_e.border = Border(left=BORDER_THIN, right=BORDER_MEDIUM, top=BORDER_THIN, bottom=bottom)


def write_holiday_row(ws: Worksheet, row: int, text, is_last: bool, tracker: ColumnWidthTracker):
    """Запись строки праздника (объединённые ячейки, красный текст)."""
    ws.merge_cells(f"C{row}:E{row}")
    cell = ws.cell(row, 3)
    text_safe = sanitize_text(text)
    cell.value = text_safe
    cell.font = FONT_HOLIDAY
    cell.alignment = ALIGN_LEFT
    cell.fill = NO_FILL
    bottom = BORDER_MEDIUM if is_last else BORDER_THIN
    cell.border = Border(left=BORDER_MEDIUM, right=BORDER_MEDIUM, top=BORDER_THIN, bottom=bottom)
    
    ws.cell(row, 4).border = Border(top=BORDER_THIN, bottom=bottom)
    ws.cell(row, 5).border = Border(right=BORDER_MEDIUM, top=BORDER_THIN, bottom=bottom)
    
    tracker.update(5, text_safe)


def fill_worksheet(
    ws: Worksheet,
    events: list[dict],
    holidays: list[dict],
    lang: str = "en",
    monday: Optional[date] = None,
) -> Optional[date]:
    """Заполнение листа данными. Возвращает дату понедельника недели."""
    tracker = ColumnWidthTracker()
    
    write_header_row(ws, lang, tracker)
    
    events_by_date = group_items_by_date(events)
    holidays_by_date = group_items_by_date(holidays)

    if monday is None:
        monday = choose_reference_monday(events_by_date, holidays_by_date)
    week_dates = get_week_dates(monday)
    
    format_date = format_date_en if lang == "en" else format_date_ru
    country_names = COUNTRY_NAMES_EN if lang == "en" else COUNTRY_NAMES_RU
    
    current_row = 3
    
    for i, d in enumerate(week_dates):
        is_first = (i == 0)
        is_last_day = (i == len(week_dates) - 1)
        
        write_date_row(ws, current_row, format_date(d), is_first=is_first, tracker=tracker)
        current_row += 1
        
        day_holidays = holidays_by_date.get(d, [])
        day_events = events_by_date.get(d, [])
        
        if day_holidays:
            holidays_grouped: dict[str, list[str]] = {}
            for hol in day_holidays:
                name = hol.get("holiday", "") or hol.get("event", "")
                country = hol.get("country", "")
                if name:
                    holidays_grouped.setdefault(name, []).append(country)
            
            holiday_parts = []
            for name, countries in holidays_grouped.items():
                country_list = ", ".join(country_names.get(c, c) for c in sorted(set(countries)))
                if lang == "en":
                    holiday_parts.append(f"{name}. Markets in {country_list}")
                else:
                    holiday_parts.append(f"{name}. Праздники в {country_list}")
            
            holiday_text = "; ".join(holiday_parts)
            is_last_row = (len(day_events) == 0) and is_last_day
            write_holiday_row(ws, current_row, holiday_text, is_last=is_last_row, tracker=tracker)
            current_row += 1
        
        day_events.sort(key=lambda x: parse_time_for_sort(x.get("time", "")))
        
        for j, ev in enumerate(day_events):
            is_last_event = (j == len(day_events) - 1) and is_last_day
            
            key_val = ev.get("Key", 0)
            if isinstance(key_val, str):
                key_val = int(key_val) if key_val.lstrip("-").isdigit() else 0
            is_important = (key_val == 1)
            
            event_text = ev.get("event", "")
            # Для русского календаря конвертируем английские месяцы в русские
            if lang == "ru":
                event_text = convert_month_suffix_to_ru(event_text)
            
            write_event_row(
                ws, current_row,
                ev.get("time", ""),
                ev.get("country", ""),
                event_text,
                is_important=is_important,
                is_last=is_last_event,
                tracker=tracker
            )
            current_row += 1
        
        if not day_holidays and not day_events:
            cell_c = ws.cell(current_row, 3)
            cell_d = ws.cell(current_row, 4)
            cell_e = ws.cell(current_row, 5)
            bottom = BORDER_MEDIUM if is_last_day else BORDER_THIN
            cell_c.border = Border(left=BORDER_MEDIUM, right=BORDER_THIN, top=BORDER_THIN, bottom=bottom)
            cell_d.border = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=bottom)
            cell_e.border = Border(left=BORDER_THIN, right=BORDER_MEDIUM, top=BORDER_THIN, bottom=bottom)
            current_row += 1
    
    tracker.apply(ws, padding=2.0)
    
    return monday


def generate_excel(work_en: list[dict], work_ru: list[dict],
                   holidays_en: list[dict], holidays_ru: list[dict]) -> BytesIO:
    """Генерация Excel файла из данных без шаблона."""
    wb = Workbook()
    
    ws_ru = wb.active
    ws_ru.title = "Календарь"
    ws_en = wb.create_sheet("Economic calendar")

    combined_events_by_date = group_items_by_date(work_en + work_ru)
    combined_holidays_by_date = group_items_by_date(holidays_en + holidays_ru)
    monday = choose_reference_monday(combined_events_by_date, combined_holidays_by_date)
    
    monday_ru = fill_worksheet(ws_ru, work_ru, holidays_ru, lang="ru", monday=monday)
    if monday_ru:
        ws_ru.title = format_sheet_name_ru(monday_ru)
    
    monday_en = fill_worksheet(ws_en, work_en, holidays_en, lang="en", monday=monday)
    if monday_en:
        ws_en.title = format_sheet_name_en(monday_en)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()
    
    return buffer
