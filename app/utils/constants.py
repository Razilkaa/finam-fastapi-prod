"""Application constants."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Цвета
COLOR_RED = "FF0000"
COLOR_BLACK = "333333"
COLOR_DATE_TEXT = "212529"

# Заливки
GRAY_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

# Выравнивание
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

# Границы
BORDER_THIN = Side(style="thin", color="000000")
BORDER_MEDIUM = Side(style="medium", color="000000")

# Шрифты
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_DATE = Font(name="Arial", size=11, bold=False, color=COLOR_DATE_TEXT)
FONT_TIME = Font(name="Arial", size=11, bold=True, color=COLOR_BLACK)
FONT_TIME_RED = Font(name="Arial", size=11, bold=True, color=COLOR_RED)
FONT_EVENT = Font(name="Arial", size=11, bold=False, color=COLOR_BLACK)
FONT_EVENT_RED = Font(name="Arial", size=11, bold=False, color=COLOR_RED)
FONT_HOLIDAY = Font(name="Arial", size=11, bold=False, color=COLOR_RED)

# Названия стран
COUNTRY_NAMES_EN = {
    "US": "US", "GB": "UK", "EU": "EU", "EA": "EU",
    "DE": "Germany", "JP": "Japan", "CN": "China", "CH": "Switzerland",
}
COUNTRY_NAMES_RU = {
    "US": "США", "GB": "Великобритания", "EU": "ЕС", "EA": "ЕС",
    "DE": "Германия", "JP": "Япония", "CN": "Китай", "CH": "Швейцария",
}

# Маппинг английских месяцев → русские (родительный падеж для "23 января")
MONTH_EN_TO_RU_GENITIVE = {
    "JAN": "января", "FEB": "февраля", "MAR": "марта", "APR": "апреля",
    "MAY": "мая", "JUN": "июня", "JUL": "июля", "AUG": "августа",
    "SEP": "сентября", "OCT": "октября", "NOV": "ноября", "DEC": "декабря",
}
# Именительный падеж для "ноябрь" (без числа)
MONTH_EN_TO_RU_NOMINATIVE = {
    "JAN": "январь", "FEB": "февраль", "MAR": "март", "APR": "апрель",
    "MAY": "май", "JUN": "июнь", "JUL": "июль", "AUG": "август",
    "SEP": "сентябрь", "OCT": "октябрь", "NOV": "ноябрь", "DEC": "декабрь",
}
# Квартал
QUARTER_EN_TO_RU = {"Q1": "1 кв.", "Q2": "2 кв.", "Q3": "3 кв.", "Q4": "4 кв."}
